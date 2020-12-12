#!/usr/bin/env python3

# DS1820 and DS18S20 have the Family Code 10
# DS18B20 has Code 28 
# DS1822 the 22.
# Don't know will this work with 'B' or x22 -sensors

import os
import sys
import openpyxl
import datetime


# Remove first comma from path if using with real sensors
dir_w1_bus = "./sys/bus/w1/devices/"

def scan_sensors():
	temp_sensors = []

	try:
		temp_sensors = os.listdir(dir_w1_bus)
	except Exception as e:
		print("*** Temperature sensor not found, check sensor connection or One-Wire settings. ***")
		raise SystemExit(e)
	else:
		# Pick only valid sensor-folders, folder starts with correct family code
		temp_sensors = [i for i in temp_sensors if i.startswith("10-")]
		if temp_sensors == []:
			raise FileNotFoundError
	
	return temp_sensors # List of temp-sensors id's
	print("from: scan_sensors()", temp_sensors) # Only for development purpose, delete after release


def read_sensors(temp_sensors):
	# Take list of temp-sensors id's as argument
	temperatures = {}

	for sensor in temp_sensors:
		try_count = 0
		check_crc = "NO"
		temp = ""

		try:
			with open(dir_w1_bus + sensor + "/w1_slave", "r") as file:
				while check_crc == "NO" and try_count <= 2:
					temp = file.read()
					# Parse crc-check, last word in line (YES/NO)
					check_crc = temp.split("\n")[0].rsplit(" ",1)[-1]
					try_count += 1
		except Exception as e:
			print(e)

		if check_crc == "YES":
			# Parse temperature, last word in line followed by 't='
			temp = temp.split("\n")[1].rsplit('t=',1)[-1]
			temp = float(temp) / 1000
			key = sensor
			value = round(temp, 1)
			temperatures[key] = value 
		else:
			temp = ""
		
		# print("from: read_sensors()", check_crc, sensor, temp) # Only for development purpose, delete after release
	
	return temperatures # Dictionary of sensor-id combined with temperature

def show_temp():
	temps = read_sensors(scan_sensors())
	for addres, temp in temps.items():
		print(addres + ":", str(temp) + "\N{DEGREE SIGN}C")

def excel_save():
	def add_temp_excel(ws_data, column_for_id, last_row, temp_for_id):
		ws_data["A" + last_row] = date_now
		ws_data["A" + last_row].number_format = 'dd.mm.yyyy h:mm'

		ws_data[column_for_id + last_row] = temp_for_id
		print("from: add_temp_excel(); ws_data+=", column_for_id + last_row, temp_for_id) # Only for development purpose, delete after release


	date_now = datetime.datetime.now()
	temps = read_sensors(scan_sensors())

	# Create initial excel file, if not exist
	excel_file = "temp_history.xlsx"
	if not os.path.isfile(excel_file):
		wb = openpyxl.Workbook() # One time excel-file initializing
		ws_data = wb.active
		ws_data.title = "data"
		ws_data['A1'] = "Day"
		ws_data['B1'] = "Time"
		ws_data['C1'] = "Temp1"
		ws_data['D1'] = "Temp2"
		ws_graph = wb.create_sheet("LastMonth")
		
		try:
			wb.save(excel_file)
		except Exception as e:
			print(e)

	# Load excel-workbook and save new data
	try:
		wb = openpyxl.load_workbook(excel_file)
	except Exception as e:
		print(e)
	else:
		ws_data = wb["data"]
		print("Ensimmäisen rivin solumäärä:", len(ws_data["1"])) # Only for development purpose, delete after release
		
		# Read excel headers for sensor-id compare
		row_headers = []
		for col in ws_data['1']:
			row_headers.append(col.value)
		print("from: excel_save(); otsikon rivinimet=", row_headers) # Only for development purpose, delete after release

		# Miltä riviltä anturi-id löytyy, jos ei ole sarakkeella vielä, lisätään se sinne
		last_row = str(ws_data.max_row + 1) # Find last row from data-worksheet
		for id in temps: # Mennään anturi-id:t lävitse
			if id in row_headers: # Tarkastetaan löytyykö id excel otsikko-riviltä
				#print("from: 'id in row_header'; id=", id, "löytyi paikalta", row_headers.index(id)) # Only for development purpose, delete after release
				column_for_id = chr(row_headers.index(id) + 97) # Muuta sarakepaikka [row_headers.index(id)] kirjaimeksi [(id:0 + ascii:97 = A)]
				#print("from: 'id in row_header; id-cell coordinates=", str(column_for_id) + str(ws_data.max_row)) # Only for development purpose, delete after release
				#print("from: 'id in row_header; id temp=", temps[id]) # Only for development purpose, delete after release
				add_temp_excel(ws_data, str(column_for_id), last_row, temps[id])
			else:
				print("from: 'id in temps' Ei löytynyt excelistä id:tä=", id) # Only for development purpose, delete after release
				print("from: 'id in row_header/else'; seuraava vapaa header-rivi=", ws_data.max_column + 1)
				column_for_id = chr(ws_data.max_column + 97) # [(id:0 + ascii:97 = A)]
				print("from: 'id in row_header/else'; coordinates=", str(column_for_id) + str(1))# Only for development purpose, delete after release
				ws_data[str(column_for_id) + str(1)] = id
				add_temp_excel(ws_data, str(column_for_id), last_row, temps[id])
		
		try:
			wb.save(excel_file)
		except Exception as e:
			print(e)

	print("from: excel_save(); sheet names:=", wb.sheetnames) # Only for development purpose, delete after release


launch_argv = []
launch_argv = sys.argv

if "-show" in launch_argv:
	show_temp()
	sys.exit()

print("from: main prog", read_sensors(scan_sensors())) # Only for development purpose, delete after release

excel_save()